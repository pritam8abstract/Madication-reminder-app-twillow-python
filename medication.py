from openpyxl import load_workbook
from datetime import datetime
from twilio.rest import Client

# Load the Excel sheet
workbook = load_workbook('medication_data.xlsx')
sheet = workbook.active

# Get current date and time
current_time = datetime.now()

# Iterate through the rows of the sheet
for row in sheet.iter_rows(min_row=2, values_only=True):
    medication_name, dosage, reminder_time = row

    # Parse the reminder time from the sheet
    reminder_time = datetime.strptime(reminder_time, '%Y-%m-%d %H:%M:')

    # Check if the reminder time is in the future
    if current_time < reminder_time:
        # Calculate the time difference in seconds
        time_difference = (reminder_time - current_time).total_seconds()

        # Create a Twilio client with your Account SID and Auth Token
        client = Client('ACa4539a5e89c89015b28ad7113f95a6b2', 'afefaa341ec8c71003d0325f1b6e10c3')

        # Create and schedule a WhatsApp message reminder
        message = client.messages.create(
            from_='whatsapp:+14155238886',  # Replace with your Twilio WhatsApp number
            body=f"Reminder: Take {dosage} of {medication_name}.",
            to='whatsapp:+919830668711',  # Replace with the recipient's WhatsApp number
            media_url='https://www.pexels.com/photo/close-up-photo-of-medicinal-drugs-159211/',  # Optionally, add media URL
            status_callback='https://example.com/callback'  # Optionally, add a callback URL
        )

        print(f"Reminder scheduled for {reminder_time}. Message SID: {message.sid}")

# Save and close the workbook
workbook.save('medication_data.xlsx')
workbook.close()
